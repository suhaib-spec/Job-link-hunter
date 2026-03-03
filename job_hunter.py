"""
Link Hunter — Jobright.ai Job Posting Monitor
==============================================
Monitors https://jobright.ai/jobs/ for new job postings and records
them in an Excel file (job_postings.xlsx).

Usage:
    1. python job_hunter.py
    2. If not logged in, SIGN IN manually in the Chrome window
    3. The script auto-detects login and starts scraping
    4. Press Ctrl+C to stop monitoring
"""

import os
import sys
import time
import signal
import re
import subprocess
import logging
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    WebDriverException,
)
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JOBS_URL = "https://jobright.ai/jobs/"
EXCEL_FILE = os.path.join(SCRIPT_DIR, "job_postings.xlsx")
CHECK_INTERVAL_MINUTES = 15
MAX_SCROLLS = 20
MAX_JOB_AGE_MINUTES = 60
SCROLL_PAUSE_SECONDS = 2.0
PAGE_LOAD_TIMEOUT = 60
CARD_LOAD_TIMEOUT = 20
LOGIN_WAIT_TIMEOUT = 300

SCRIPT_CHROME_PROFILE = os.path.join(SCRIPT_DIR, ".chrome_profile")

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("LinkHunter")

_shutdown = False

def _signal_handler(sig, frame):
    global _shutdown
    log.info("Shutdown requested...")
    _shutdown = True

signal.signal(signal.SIGINT, _signal_handler)
signal.signal(signal.SIGBREAK, _signal_handler)


# ═══════════════════════════════════════════════════════════════════════════
# Chrome / Selenium
# ═══════════════════════════════════════════════════════════════════════════

def _is_first_run() -> bool:
    return not os.path.exists(os.path.join(SCRIPT_CHROME_PROFILE, "Default", "Preferences"))

def _kill_stale_chrome():
    try:
        lock_file = os.path.join(SCRIPT_CHROME_PROFILE, "lockfile")
        if os.path.exists(lock_file):
            log.info("Cleaning up stale Chrome...")
            subprocess.run(['taskkill', '/F', '/IM', 'chromedriver.exe'],
                           capture_output=True, timeout=5)
            time.sleep(1)
            try:
                os.remove(lock_file)
            except Exception:
                pass
    except Exception:
        pass

def create_driver() -> webdriver.Chrome:
    _kill_stale_chrome()
    first_run = _is_first_run()
    opts = Options()
    opts.add_argument(f"--user-data-dir={SCRIPT_CHROME_PROFILE}")
    opts.add_argument("--profile-directory=Default")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--remote-debugging-port=0")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--start-maximized")
    opts.add_argument("--no-first-run")
    opts.add_argument("--no-default-browser-check")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--log-level=3")
    opts.add_argument("--silent")

    if first_run:
        log.info("=" * 60)
        log.info("  FIRST RUN - Chrome will open. Log in to jobright.ai!")
        log.info("=" * 60)

    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=opts)
    except WebDriverException as e:
        log.error(f"Failed to start Chrome: {e}")
        log.error("  1. Close any stuck chrome.exe in Task Manager")
        log.error("  2. Delete .chrome_profile folder and retry")
        sys.exit(1)

    driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    return driver


# ═══════════════════════════════════════════════════════════════════════════
# Login detection
# ═══════════════════════════════════════════════════════════════════════════

def _is_on_jobs_page(driver: webdriver.Chrome) -> bool:
    url = driver.current_url.lower()
    if "/jobs/" in url:
        try:
            driver.find_element(By.ID, "scrollableDiv")
            return True
        except NoSuchElementException:
            pass
    return False

def wait_for_login(driver: webdriver.Chrome) -> bool:
    log.info(f"Navigating to {JOBS_URL}")
    try:
        driver.get(JOBS_URL)
    except TimeoutException:
        pass
    time.sleep(4)

    if _is_on_jobs_page(driver):
        log.info("Already logged in!")
        return True

    log.info("=" * 60)
    log.info("  LOGIN REQUIRED - Sign in in the browser window.")
    log.info(f"  (Waiting up to {LOGIN_WAIT_TIMEOUT // 60} minutes)")
    log.info("=" * 60)

    start = time.time()
    while time.time() - start < LOGIN_WAIT_TIMEOUT:
        if _shutdown:
            return False
        time.sleep(3)
        try:
            if _is_on_jobs_page(driver):
                log.info("Login detected!")
                return True
        except WebDriverException:
            pass

    log.error("Timed out waiting for login.")
    return False


# ═══════════════════════════════════════════════════════════════════════════
# Sort by "Most Recent" — uses Selenium ActionChains for reliable clicking
# ═══════════════════════════════════════════════════════════════════════════

def _select_most_recent(driver: webdriver.Chrome) -> bool:
    """Click the sort dropdown and select 'Most Recent'."""
    log.info("Selecting 'Most Recent' sort order...")

    try:
        # Find the sort trigger using XPath — it contains text like
        # "Recommended", "Top Matched", or "Most Recent"
        # There are TWO "Recommended" elements: one is the nav tab, the other
        # is the sort dropdown. The sort dropdown is AFTER the help "?" icon.
        sort_trigger = None

        # Strategy 1: Find by XPath - elements with exact sort text that are
        # NOT inside the navigation tabs (the nav tab has a different style)
        for text in ["Recommended", "Most Recent", "Top Matched"]:
            try:
                elements = driver.find_elements(
                    By.XPATH,
                    f"//*[normalize-space(text())='{text}']"
                )
                for el in elements:
                    try:
                        # The sort dropdown version is typically:
                        # - In the filter area (y between 80-180px)
                        # - Has a parent/sibling with a dropdown arrow
                        # - Is NOT the main nav tab
                        rect = el.rect
                        parent_class = el.find_element(By.XPATH, "..").get_attribute("class") or ""

                        # Skip the nav tab (it's typically a link or has specific classes)
                        tag = el.tag_name.lower()
                        if tag == "a" or "nav" in parent_class.lower():
                            continue

                        # The sort dropdown is to the right (x > 40% of viewport)
                        viewport_width = driver.execute_script("return window.innerWidth;")
                        if rect["x"] > viewport_width * 0.4:
                            if text == "Most Recent":
                                log.info("  Already sorted by 'Most Recent'.")
                                return True
                            sort_trigger = el
                            break
                    except Exception:
                        continue
                if sort_trigger:
                    break
            except Exception:
                continue

        # Strategy 2: Find any element with class containing "filter" and sort text
        if not sort_trigger:
            try:
                candidates = driver.find_elements(
                    By.CSS_SELECTOR,
                    "[class*='filter'] span, [class*='sort'] span, [class*='dropdown-trigger']"
                )
                for c in candidates:
                    try:
                        t = c.text.strip().lower()
                        if t in ("recommended", "top matched", "most recent"):
                            if t == "most recent":
                                log.info("  Already sorted by 'Most Recent'.")
                                return True
                            sort_trigger = c
                            break
                    except StaleElementReferenceException:
                        continue
            except Exception:
                pass

        if not sort_trigger:
            log.warning("  Sort dropdown trigger not found.")
            return False

        # Click the trigger using ActionChains (more reliable than .click())
        log.info("  Clicking sort dropdown...")
        ActionChains(driver).move_to_element(sort_trigger).click().perform()
        time.sleep(2)

        # Now find and click "Most Recent" in the opened menu
        # Try multiple times since the menu renders asynchronously
        for attempt in range(6):
            try:
                # Look for menu items containing "Most Recent"
                menu_items = driver.find_elements(
                    By.XPATH,
                    "//*[normalize-space(text())='Most Recent']"
                )
                for item in menu_items:
                    try:
                        if item.is_displayed():
                            ActionChains(driver).move_to_element(item).click().perform()
                            log.info(f"  Selected 'Most Recent'! (attempt {attempt + 1})")
                            time.sleep(3)
                            return True
                    except Exception:
                        continue
            except Exception:
                pass
            time.sleep(0.5)

        log.warning("  Could not click 'Most Recent' in the menu.")
        ActionChains(driver).move_by_offset(0, 0).click().perform()
        return False

    except Exception as e:
        log.warning(f"  Dropdown error: {e}")
        return False


# ═══════════════════════════════════════════════════════════════════════════
# Job extraction — handles VIRTUALIZED scrolling
# ═══════════════════════════════════════════════════════════════════════════

# Extracts CURRENTLY VISIBLE cards. Called repeatedly during scrolling.
# Also clicks each card's title to capture the job detail URL.
JS_EXTRACT_VISIBLE = r"""
return (function() {
    const container = document.getElementById('scrollableDiv');
    if (!container) return [];

    const h2s = container.querySelectorAll('h2');
    const jobs = [];

    for (const h2 of h2s) {
        try {
            const title = h2.textContent.trim();
            if (!title || title.length < 3) continue;

            // Walk up to find card container
            let card = h2.parentElement;
            for (let i = 0; i < 8 && card && card !== container; i++) {
                const r = card.getBoundingClientRect();
                if (r.height > 120 && r.width > 300) break;
                card = card.parentElement;
            }
            if (!card || card === container) card = h2.parentElement;

            const text = (card.innerText || '').trim();
            const lines = text.split('\n').map(l => l.trim()).filter(l => l.length > 0);

            // Company
            let company = '';
            for (let i = 0; i < lines.length; i++) {
                if (lines[i] === title && i + 1 < lines.length) {
                    const next = lines[i + 1];
                    if (next && !next.includes('$') &&
                        !next.toLowerCase().includes('ago') &&
                        !next.toLowerCase().includes('applicant')) {
                        company = next.split('/')[0].trim();
                        break;
                    }
                }
            }

            // Spans for metadata
            const spans = Array.from(card.querySelectorAll('span'))
                .map(s => s.textContent.trim())
                .filter(t => t.length > 0 && t.length < 100);

            let datePosted = '', location = '', salary = '';
            let jobType = '', workModel = '';

            for (const s of spans) {
                const low = s.toLowerCase();
                if ((low.includes('ago') || low.includes('just now')) && !datePosted)
                    datePosted = s;
                else if (s.includes('$') && !salary)
                    salary = s;
                else if (['full-time','part-time','contract','internship','temporary'].includes(low))
                    jobType = s;
                else if (['remote','onsite','on-site','hybrid'].includes(low))
                    workModel = s;
            }

            for (const s of spans) {
                if (!location && /[A-Z][a-z]+,\s*[A-Z]{2}/.test(s)) { location = s; break; }
            }
            if (!location) {
                for (const s of spans) {
                    if (s.includes('United States') || s === 'Remote') { location = s; break; }
                }
            }

            if (!datePosted) {
                for (const l of lines) {
                    if (l.toLowerCase().includes('ago') || l.toLowerCase().includes('just now')) {
                        datePosted = l; break;
                    }
                }
            }

            // Try to find job URL from any link in/around the card
            let jobUrl = '';
            const links = card.querySelectorAll('a[href]');
            for (const a of links) {
                if (a.href.includes('/jobs/info/') || a.href.includes('/jobs/')) {
                    jobUrl = a.href;
                    break;
                }
            }

            // Try APPLY NOW button - sometimes it's a link
            if (!jobUrl) {
                const applyBtns = card.querySelectorAll('a[href], button[onclick]');
                for (const btn of applyBtns) {
                    const href = btn.getAttribute('href') || '';
                    if (href && href.includes('job')) {
                        jobUrl = href.startsWith('http') ? href : 'https://jobright.ai' + href;
                        break;
                    }
                }
            }

            // Try data attributes on card or ancestors
            if (!jobUrl) {
                let el = card;
                for (let i = 0; i < 5 && el; i++) {
                    const attrs = el.attributes;
                    for (let a = 0; a < attrs.length; a++) {
                        const val = attrs[a].value;
                        if (val && /^[a-f0-9]{24}$/i.test(val)) {
                            jobUrl = 'https://jobright.ai/jobs/info/' + val;
                            break;
                        }
                    }
                    if (jobUrl) break;
                    // Check ID attribute
                    if (el.id && /^[a-f0-9]{24}$/i.test(el.id)) {
                        jobUrl = 'https://jobright.ai/jobs/info/' + el.id;
                        break;
                    }
                    el = el.parentElement;
                }
            }

            // Generate stable ID for dedup
            let hash = 0;
            const str = (title + '|' + company).toLowerCase().replace(/\s+/g, '');
            for (let c = 0; c < str.length; c++) {
                hash = ((hash << 5) - hash) + str.charCodeAt(c);
                hash = hash & hash;
            }
            const jobId = 'jh_' + Math.abs(hash).toString(16);

            jobs.push({
                job_id: jobId,
                title: title,
                company: company,
                location: location,
                salary: salary,
                job_type: jobType,
                work_model: workModel,
                date_posted: datePosted,
                job_url: jobUrl
            });
        } catch(e) {
            continue;
        }
    }
    return jobs;
})();
"""


def _is_within_time_limit(date_text: str) -> bool:
    if not date_text:
        return True
    lower = date_text.lower().strip()
    if "just now" in lower or "second" in lower:
        return True
    m = re.search(r'(\d+)\s*minute', lower)
    if m:
        return int(m.group(1)) <= MAX_JOB_AGE_MINUTES
    m = re.search(r'(\d+)\s*hour', lower)
    if m:
        return (int(m.group(1)) * 60) <= MAX_JOB_AGE_MINUTES
    if "day" in lower or "week" in lower or "month" in lower or "year" in lower:
        return False
    return True


def _try_get_job_urls(driver: webdriver.Chrome, jobs: list[dict]):
    """
    For jobs without URLs, try clicking each card title in the browser
    to capture the job detail URL from the page navigation.
    """
    jobs_needing_urls = [j for j in jobs if not j.get("job_url")]
    if not jobs_needing_urls:
        return

    log.info(f"  Attempting to find URLs for {len(jobs_needing_urls)} jobs...")

    for j in jobs_needing_urls[:10]:  # Limit to avoid being too slow
        if _shutdown:
            break
        try:
            # Find the h2 with this exact title and click it
            h2s = driver.find_elements(By.CSS_SELECTOR, "#scrollableDiv h2")
            for h2 in h2s:
                try:
                    if h2.text.strip() == j["title"]:
                        ActionChains(driver).move_to_element(h2).click().perform()
                        time.sleep(1.5)

                        # Check if URL changed to include a job ID
                        current = driver.current_url
                        if "/jobs/info/" in current:
                            j["job_url"] = current.split("?")[0]
                            log.info(f"    Got URL for: {j['title']}")
                        break
                except StaleElementReferenceException:
                    continue
        except Exception:
            continue

    # Navigate back to jobs page if we navigated away
    if "/jobs/info/" in driver.current_url:
        try:
            driver.get(JOBS_URL)
            time.sleep(2)
        except Exception:
            pass


def scrape_jobs(driver: webdriver.Chrome) -> list[dict]:
    """Scrape job postings with incremental extraction during scrolling."""

    if not _is_on_jobs_page(driver):
        if not wait_for_login(driver):
            return []

    log.info("Refreshing jobs page...")
    try:
        driver.get(JOBS_URL)
    except TimeoutException:
        pass

    try:
        WebDriverWait(driver, CARD_LOAD_TIMEOUT).until(
            EC.presence_of_element_located((By.ID, "scrollableDiv"))
        )
    except TimeoutException:
        log.warning("Job list not found.")
        driver.save_screenshot(os.path.join(SCRIPT_DIR, "debug_screenshot.png"))
        return []

    time.sleep(3)

    # Select "Most Recent" sort
    _select_most_recent(driver)

    # Incremental scroll + extract
    all_jobs = {}
    no_new_count = 0

    log.info("Scrolling and extracting jobs...")

    # Extract initially visible cards
    try:
        batch = driver.execute_script(JS_EXTRACT_VISIBLE)
        if batch:
            for j in batch:
                jid = j.get("job_id", "")
                if jid and jid not in all_jobs:
                    all_jobs[jid] = j
            log.info(f"  Initial: {len(batch)} visible, {len(all_jobs)} unique")
    except Exception as e:
        log.warning(f"  Initial extraction error: {e}")

    for i in range(MAX_SCROLLS):
        if _shutdown:
            break

        try:
            driver.execute_script(
                "const el = document.getElementById('scrollableDiv');"
                "if (el) el.scrollTop += el.clientHeight * 0.8;"
            )
        except Exception:
            break

        time.sleep(SCROLL_PAUSE_SECONDS)

        try:
            batch = driver.execute_script(JS_EXTRACT_VISIBLE)
        except Exception:
            continue

        if not batch:
            no_new_count += 1
            if no_new_count >= 3:
                log.info(f"  No cards for 3 scrolls — end reached.")
                break
            continue

        new_in_batch = 0
        for j in batch:
            jid = j.get("job_id", "")
            if jid and jid not in all_jobs:
                all_jobs[jid] = j
                new_in_batch += 1

        if new_in_batch > 0:
            no_new_count = 0
            log.info(f"  Scroll {i+1}: +{new_in_batch} new ({len(all_jobs)} total)")
        else:
            no_new_count += 1
            if no_new_count >= 3:
                log.info(f"  No new jobs for 3 scrolls — done at {len(all_jobs)} total.")
                break

    # Filter by recency
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    jobs = []
    for j in all_jobs.values():
        j["scraped_at"] = now
        if _is_within_time_limit(j.get("date_posted", "")):
            jobs.append(j)

    log.info(f"Total unique: {len(all_jobs)}, within last {MAX_JOB_AGE_MINUTES} min: {len(jobs)}")

    # Try to get real job URLs for jobs that don't have them
    if jobs:
        _try_get_job_urls(driver, jobs)
        s = jobs[0]
        log.info(f"  Sample: \"{s['title']}\" @ {s['company']} -- {s['date_posted']}")
        if s.get("job_url"):
            log.info(f"  URL: {s['job_url']}")

    return jobs


# ═══════════════════════════════════════════════════════════════════════════
# Excel management
# ═══════════════════════════════════════════════════════════════════════════

HEADERS = [
    "Job Link", "Title", "Company", "Location", "Salary",
    "Job Type", "Work Model", "Date Posted", "First Seen",
]

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NEW_ROW_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")


def _create_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Postings"
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    widths = [55, 50, 30, 25, 22, 14, 12, 18, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(HEADERS))}1"
    return wb


def load_existing_ids() -> set[str]:
    if not os.path.exists(EXCEL_FILE):
        return set()
    try:
        wb = load_workbook(EXCEL_FILE, read_only=True)
        ws = wb.active
        ids = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                ids.add(str(row[0]))
        wb.close()
        return ids
    except Exception as e:
        log.warning(f"Could not read Excel: {e}")
        return set()


def save_new_jobs(jobs: list[dict]) -> int:
    existing_ids = load_existing_ids()
    # Use URL as key if available, otherwise the hash ID
    new_jobs = []
    for j in jobs:
        key = j.get("job_url") or j["job_id"]
        if key and key not in existing_ids:
            new_jobs.append(j)

    if not new_jobs:
        log.info("No new jobs to add.")
        return 0

    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE)
        except Exception:
            wb = _create_workbook()
    else:
        wb = _create_workbook()

    ws = wb.active
    next_row = ws.max_row + 1
    for job in new_jobs:
        row_data = [
            job.get("job_url") or job["job_id"],
            job["title"], job["company"], job["location"],
            job["salary"], job["job_type"], job["work_model"],
            job["date_posted"], job["scraped_at"],
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.fill = NEW_ROW_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        next_row += 1

    # Save with retry — keeps trying the main file, falls back to alternate
    temp_path = EXCEL_FILE + ".tmp"
    for attempt in range(3):
        try:
            wb.save(temp_path)
            wb.close()
            if os.path.exists(EXCEL_FILE):
                os.replace(temp_path, EXCEL_FILE)
            else:
                os.rename(temp_path, EXCEL_FILE)
            log.info(f"Added {len(new_jobs)} new job(s) to {os.path.basename(EXCEL_FILE)}")
            return len(new_jobs)
        except PermissionError:
            if attempt < 2:
                log.warning(f"  Excel file is open! Please close it. Retrying in 10s... ({attempt+1}/3)")
                time.sleep(10)
            else:
                alt_file = os.path.join(SCRIPT_DIR, f"job_postings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                try:
                    wb.save(alt_file)
                    wb.close()
                    log.warning(f"  Could not write to main file. Saved to: {os.path.basename(alt_file)}")
                    return len(new_jobs)
                except Exception as e2:
                    log.error(f"  Could not save to alternate file either: {e2}")
        except Exception as e:
            log.error(f"Failed to save Excel: {e}")
            break

    if os.path.exists(temp_path):
        try:
            os.remove(temp_path)
        except Exception:
            pass
    return 0


# ═══════════════════════════════════════════════════════════════════════════
# Main loop
# ═══════════════════════════════════════════════════════════════════════════

def print_banner():
    print()
    print("  ======================================================")
    print("  ||    Link Hunter  --  Job Posting Monitor           ||")
    print("  ||    Tracking: jobright.ai/jobs/ (Most Recent)      ||")
    print("  ======================================================")
    print()
    print("  A Chrome window will open for scraping.")
    print("  On first run, log in to jobright.ai.")
    print("  IMPORTANT: Close the Excel file before running!")
    print()


def main():
    print_banner()
    log.info(f"Excel: {EXCEL_FILE}")
    log.info(f"Interval: {CHECK_INTERVAL_MINUTES} min | Max age: {MAX_JOB_AGE_MINUTES} min")
    log.info("Starting Chrome...")

    driver = create_driver()
    cycle = 0

    try:
        log.info("Checking login status...")
        if not wait_for_login(driver):
            log.error("Could not access jobs page. Exiting.")
            return

        while not _shutdown:
            cycle += 1
            log.info(f"{'=' * 55}")
            log.info(f"  Cycle #{cycle} -- {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            log.info(f"{'=' * 55}")

            try:
                jobs = scrape_jobs(driver)
                if jobs:
                    new_count = save_new_jobs(jobs)
                    total = len(load_existing_ids())
                    log.info(f"Summary: {len(jobs)} scraped | {new_count} new | {total} total")
                else:
                    log.warning("No recent jobs found this cycle.")
            except Exception as e:
                log.error(f"Error during scrape: {e}")
                try:
                    driver.quit()
                except Exception:
                    pass
                log.info("Restarting browser...")
                driver = create_driver()
                wait_for_login(driver)

            if _shutdown:
                break

            log.info(f"Next check in {CHECK_INTERVAL_MINUTES} min. Ctrl+C to stop.")
            for _ in range(CHECK_INTERVAL_MINUTES * 60):
                if _shutdown:
                    break
                time.sleep(1)

    finally:
        log.info("Shutting down browser...")
        try:
            driver.quit()
        except Exception:
            pass
        log.info("Link Hunter stopped. Goodbye!")


if __name__ == "__main__":
    main()
